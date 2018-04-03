# https://hackernoon.com/raschietto-a-simple-library-for-web-scraping-46957c6aa5b7

from raschietto import Raschietto, Matcher
import pandas as pd
from openpyxl import load_workbook, Workbook
from progress.bar import Bar
import os
from sys import platform

def update(url):
    # url = 'https://www.robotevents.com/robot-competitions/vex-robotics-competition/RE-VRC-17-2609.html'
    # url = 'https://www.robotevents.com/robot-competitions/vex-robotics-competition/RE-VRC-17-3805.html'
    # url = 'https://www.robotevents.com/robot-competitions/vex-robotics-competition/RE-VRC-17-4462.html'
    page = Raschietto.from_url(url)

    with open('data/urls.txt', 'w+') as filehandle:
        filehandle.write(url)

    team = Matcher('.table-bordered>tbody>tr>td')
    teams = team(page, multiple=True)

    teamNums = []
    teamNames = []
    organizations = []
    locations = []

    for i in range(len(teams)):
        if (i + 1) % 4 == 0:
            locations.append(teams[i].replace('\n', ' '))
            organizations.append(teams[i - 1])
            teamNames.append(teams[i - 2])
            teamNums.append(teams[i - 3])


    allEvents = []
    allRanks = []
    allWLTs = []
    allWPSPs = []
    allMAXSCORES = []
    allOPRS = []

    # teamEventsDict = {}
    # teamRanksDict = {}
    # teamWLTsDict = {}
    # teamWPSPsDict = {}
    # teamMAXSCORESDict = {}
    # teamOPRSDict = {}
    bar = Bar('Collecting Data', max=len(teamNums))

    wb = Workbook()
    wb.create_sheet('teamlist')
    # print(teamNums)
    for number in teamNums:
        wb.create_sheet(number)
    writer = pd.ExcelWriter("data/data.xlsx")

    for teamNumber in teamNums:
        scores = Raschietto.from_url("https://vexdb.io/teams/view/{}?t=rankings".format(teamNumber))
        eventPointer = Matcher('.event')
        eventsList = eventPointer(scores, multiple=True)
        try: allEvents.append(eventsList[1])
        except: allEvents.append("New Team")
        try: eventsList.remove(eventsList[0])
        except: pass
        # teamEventsDict[teamNumber] = eventsList

        rankPointer = Matcher('.rank')
        ranksList = rankPointer(scores, multiple=True)
        try: allRanks.append(ranksList[1])
        except: allRanks.append("New Team")
        try: ranksList.remove(ranksList[0])
        except: pass
        ranksList = list(map(int, ranksList))
        # teamRanksDict[teamNumber] = ranksList

        wltPointer = Matcher('.wlt')
        wltList = wltPointer(scores, multiple=True)
        try: allWLTs.append(wltList[1])
        except: allWLTs.append("New Team")
        try: wltList.remove(wltList[0])
        except: pass
        # teamWLTsDict[teamNumber] = eventsList

        WPSPPointer = Matcher('.wpsp')
        WPSPList = WPSPPointer(scores, multiple=True)
        try: allWPSPs.append(WPSPList[1])
        except: allWPSPs.append("New Team")
        try: WPSPList.remove(WPSPList[0])
        except: pass
        # teamWPSPsDict[teamNumber] = WPSPList

        maxScorePointer = Matcher('.max_score')
        maxScoreList = maxScorePointer(scores, multiple=True)
        try: allMAXSCORES.append(maxScoreList[1])
        except: allMAXSCORES.append("New Team")
        try: maxScoreList.remove(maxScoreList[0])
        except: pass
        maxScoreList = list(map(int, maxScoreList))

        # teamMAXSCORESDict[teamNumber] = maxScoreList

        OPRPointer = Matcher('.opr')
        OPRList = OPRPointer(scores, multiple=True)
        try: allOPRS.append(OPRList[1])
        except: allOPRS.append("New Team")
        try: OPRList.remove(OPRList[0])
        except: pass
        OPRList = list(map(float, OPRList))

        # teamOPRSDict[teamNumber] = OPRList


        pdTeamEvents = pd.Series(data=eventsList)
        pdTeamRanks = pd.Series(data=ranksList)
        pdTeamWLTs = pd.Series(data=wltList)
        pdTeamWPSPs = pd.Series(data=WPSPList)
        pdTeamMAXSCORES = pd.Series(data=maxScoreList)
        pdTeamOPRS = pd.Series(data=OPRList)
        pdTeamStats = pd.DataFrame({"Events":pdTeamEvents, "Ranks":pdTeamRanks, "WLT":pdTeamWLTs, "WPSP":pdTeamWPSPs, "Max Score":pdTeamMAXSCORES,"OPR":pdTeamOPRS})
        pdTeamStats.to_excel(writer, sheet_name=teamNumber)
        pdTeamStats.to_pickle("team_dataframes/"+teamNumber+".pkl")
        bar.next()

    pdEvents = pd.Series(data=allEvents)
    pdRanks = pd.Series(data=allRanks)
    pdWLT = pd.Series(data=allWLTs)
    pdWPSP = pd.Series(data=allWPSPs)
    pdMaxScore = pd.Series(data=allMAXSCORES)
    pdOPR = pd.Series(data=allOPRS)

    pdTeamNums = pd.Series(data=teamNums)
    pdTeamNames = pd.Series(data=teamNames)
    pdOrganizations = pd.Series(data=organizations)
    pdLocations = pd.Series(data=locations)

    teamList = pd.DataFrame({ "Team Number: " : pdTeamNums, "Team Name: ": pdTeamNames, "Organization: ": pdOrganizations, "Location: ": pdLocations, "Most Recent Event: ": pdEvents, "Rank: ": pdRanks, "W-L-T: ": pdWLT, "WPSP: ": pdWPSP, "Max Score: ": pdMaxScore, "OPR: ": pdOPR})
    teamList = teamList[["Team Number: ", "Team Name: ", "Organization: ", "Location: ", "Most Recent Event: ", "Rank: ", "W-L-T: ", "WPSP: ", "Max Score: ", "OPR: "]]

    teamList.to_excel(writer,'teamlist')
    teamList.to_pickle("team_dataframes/teamList.pkl")

# df = pd.read_pickle(file_name)

    wb.save("data/data.xlsx")

    writer.save()

    wb = load_workbook("data/data.xlsx")
    sheet = wb.get_sheet_by_name('teamlist')
    sheet.column_dimensions['B'].width = len(max(teamNums)) * 1.2
    sheet.column_dimensions['C'].width = len(max(teamNames)) * 1.2
    sheet.column_dimensions['D'].width = len(max(organizations)) * 1.2
    sheet.column_dimensions['E'].width = len(max(locations)) * 0.8
    sheet.column_dimensions['F'].width = 40 # len(max(allEvents))
    sheet.column_dimensions['G'].width = 6 # len(max(allRanks)) * 1.2
    sheet.column_dimensions['H'].width = len(max(allWLTs)) * 1.2
    sheet.column_dimensions['I'].width = len(max(allWPSPs)) * 1.2
    sheet.column_dimensions['J'].width = 10 #len(max(allMAXSCORES)) * 1.2
    sheet.column_dimensions['K'].width = len(max(allOPRS)) * 1.2
    wb.save("data/data.xlsx")

    bar.finish()
    print("Completed")
    # For Mac
    if platform == "darwin":
        try: os.system("open -a 'Microsoft Excel.app' 'data/data.xlsx'")
        except: print("Failed to open file. Please open it on your own.")
    elif platform == "win32":
        try: os.system("open -a 'Microsoft Excel.exe' 'data/data.xlsx'")
        except: print("Failed to open file. Please open it on your own.")
